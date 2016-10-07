'''
Convert a row/col coordinate to an excel 'D5' cell location
'''
def getCellName(row, col):
	alpha = ''
	while col > 0:
		d = int(col % 26)
		if d == 0:
			d = 26
		col -= d
		alpha = str(chr(65+d-1)) + alpha
		col = col / 26
	return alpha + str(row)

'''
Parse persons and their preferences off of one Excel worksheet
'''
def grabListsFromSheet(filepath):
	from openpyxl import load_workbook
	sheet = load_workbook(filepath).active
	persons = []
	prefs = {}
	row = 2
	while sheet.cell(row=row, column=2).value != None:
		person = sheet.cell(row=row, column=2).value
		persons.append(person)
		prefs[person] = []
		col = 3
		while sheet.cell(row=row, column=col).value != None:
			prefs[person].append(sheet.cell(row=row, column=col).value)
			col += 1
		row += 1
	return persons, prefs

'''
Parse persons and their preferences off of one .csv file
'''
def grabListsFromCSV(filepath):
	import csv
	csvfile = open(filepath, 'r')
	reader = csv.reader(csvfile)
	persons = []
	prefs = {}
	cnt = 1
	for row in reader:
		if cnt == 1:
			cnt += 1
			continue
		persons.append(row[1])
		prefs[row[1]] = row[2:]
	return persons, prefs

'''
Duplicate people to ensure there are an equal number of bigs and littles
'''
def completePersons(bigs, littles, bigsPreferences, littlesPreferences):
	cnt = 0
	while len(bigs) < len(littles):
		bigs.append(bigs[cnt] + '*')
		bigsPreferences[bigs[cnt] + '*'] = bigsPreferences[bigs[cnt]]
		cnt += 1
	while len(littles) < len(bigs):
		littles.append(littles[cnt] + '*')
		littlesPreferences[littles[cnt] + '*'] = littlesPreferences[littles[cnt]]
		cnt += 1
	return bigs, littles, bigsPreferences, littlesPreferences

'''
Fill in preferences with the rest of potential candidates
A person is likely indifferent about these filled-in candidates as they were explicitly specificied in their preferences
'''
def completePreferences(prefs, prefValues):
	for person in prefs:
		for prefValue in prefValues:
			if prefValue not in prefs[person]:
				prefs[person].append(prefValue)
	return prefs

'''
Create lists of bigs and littles and dictionaries for their respective preferences
'''
def initializeLists(bigsPath, littlesPath):
	if bigsPath[bigsPath.rfind('.'):] == '.xlsx':
		bigs, bigsPreferences = grabListsFromSheet(bigsPath)
	elif bigsPath[bigsPath.rfind('.'):] == '.csv':
		bigs, bigsPreferences = grabListsFromCSV(bigsPath)
	if littlesPath[littlesPath.rfind('.'):] == '.xlsx':
		littles, littlesPreferences = grabListsFromSheet(littlesPath)
	elif littlesPath[littlesPath.rfind('.'):] == '.csv':
		littles, littlesPreferences = grabListsFromCSV(littlesPath)
	bigs, littles, bigsPreferences, littlesPreferences = completePersons(bigs, littles, bigsPreferences, littlesPreferences)
	bigsPreferences = completePreferences(bigsPreferences, littles)
	littlesPreferences = completePreferences(littlesPreferences, bigs)
	return bigs, littles, bigsPreferences, littlesPreferences

'''
Initialize dictionaries with the list as the keys and boolean Falses as the values
'''
def convertListToDict(list):
	toDict = {}
	for el in list:
		toDict[el] = False
	return toDict

'''
Return the pair with the person
'''
def pairWith(pairs, person):
	for pair in pairs:
		if person in pair:
			return pair
	assert 'Pair not found'

'''
Use the Gale-Shapley algorithm to pair up bigs and littles based upon their preferences
A big and a little cannot be paired with another respective little or big if they prefer each other.
'''
def pairBigLittles(bigs, littles, bigsPreferences, littlesPreferences):
	pairs = []
	bigsDict = convertListToDict(bigs)
	littlesDict = convertListToDict(littles)
	#Bigs propose to their top little
	for big in bigsDict:
		little = bigsPreferences[big].pop(0)
		#Littles only accept proposal if from top bigs
		if big == littlesPreferences[little][0]:
			pairs.append([big, little])
			bigsDict[big] = True
			littlesDict[little] = True
	#while any bigs don't have a little
	while False in bigsDict.values():
		for big in bigsDict:
			#skip bigs that already have littles
			if bigsDict[big]:
				continue
			little = bigsPreferences[big].pop(0)
			#if little does not have a big, pair big/little
			if not littlesDict[little]:
				bigsDict[big] = True
				littlesDict[little] = True
				pairs.append([big, little])
			#if little does have a big but prefers this new big, remove old pair and pair big/little
			else:
				otherPair = pairWith(pairs, little)
				if littlesPreferences[little].index(big) < littlesPreferences[little].index(otherPair[0]):
					bigsDict[otherPair[0]] = False
					bigsDict[big] = True
					littlesDict[little] = True #should be redundant
					pairs.remove(otherPair)
					pairs.append([big, little])
	return pairs

'''
Print the pairs into the same workbook
'''
def printPairsToWorkbook(pairs):
	from openpyxl import Workbook
	wb = Workbook()
	sheet = wb.active
	sheet.title = 'Pairs'
	sheet['A1'] = 'Bigs'
	sheet['B1'] = 'Littles'
	row = 2
	for pair in pairs:
		sheet.cell(row=row, column=1).value = pair[0]
		sheet.cell(row=row, column=2).value = pair[1]
		row += 1
	wb.save('pairs.xlsx')

def main():
	bigs, littles, bigsPreferences, littlesPreferences = initializeLists('bigs.xlsx','littles.csv')
	pairs = pairBigLittles(bigs, littles, bigsPreferences, littlesPreferences)
	printPairsToWorkbook(pairs)
main()